import json
import os
from bson import ObjectId
from flask import Blueprint, url_for, session, request, render_template, redirect, g, flash, jsonify
from persiantools import digits
from pymongo import MongoClient
import io
from json import load

from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
import openpyxl
from pathlib import Path
from datetime import datetime
from persiantools.jdatetime import JalaliDateTime, JalaliDate

bp = Blueprint('api', __name__, url_prefix="/api")


def return_category(filename):
    li = []
    with io.open(filename, encoding='utf-8') as f:
        f = json.load(f)
        category = f[0]
        for i in range(len(category['subcategories'])):
            cat2 = category['name']
            cat2 += r'/' + category['subcategories'][i]['name']
            for j in range(len(category['subcategories'][0]['subcategoreis'])):
                cat3 = cat2
                cat3 += r'/' + category['subcategories'][i]['subcategoreis'][j]['name']
                li.append(cat3)
                cat3 = cat2
    return li


def check_validation_excel(path):
    category = []
    xlsx_file = Path(path)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    for row in sheet.iter_rows(max_row=sheet.max_row):
        if row[-2].value not in return_category(r'onlineshop/category.json'):
            raise Exception('your category is wrong')
        else:
            category.append({
                'name_product': row[-1].value,
                'category': row[-2].value,
                'descrption': row[-3].value,
                'url_image': row[-4].value
            })
    return category


def update(db, request):
    res = db.inventory.update_many(
        {"_id": ObjectId(request.form.get('_id')), "items.name_product": request.form.get('name_product')},
        {"$inc": {"items.$.quantity": int(request.form.get('quantity'))},
         "$set": {"items.$.price": int(request.form.get('price')), "items.$.date_insert": datetime.now()}})
    if res.modified_count == 0:
        return True
    else:
        return False


def update_reverse(db, request):
    res = db.inventory.update_many(
        {"_id": ObjectId(request.form.get('_id')), "items.name_product": request.form.get('name_product')},
        {"$inc": {"items.$.quantity": -int(request.form.get('quantity'))}})
    if res.modified_count == 0:
        return True
    else:
        return False


def has_image(f):
    if len(f) == 0:
        return '/static/images/surface.jpg'
    else:
        f['file'].save(r'onlineshop/static/images/' + secure_filename(f['file'].filename))
        return fr"/static/images/{secure_filename(f['file'].filename)}"


@bp.route('/product/list')
def prod_list():
    client = MongoClient('localhost', 27017)
    db = client.online_shop
    prod_list = list(db.products.find())
    for i in prod_list:
        i["_id"] = str(i["_id"])
    return jsonify(prod_list)


@bp.route('product/return_product_id', methods=('GET', 'POST'))
def return_product_id():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        res = list(db.products.find({"name_product": request.form["name_product"]}, {"_id": 1}))
        res = [{"_id": str(i["_id"])} for i in res]
        if len(res) > 0:
            return jsonify(res)
        else:
            return jsonify([{"_id": "#"}])


@bp.route('/product/<product_id>')
def prod_details(product_id, db):
    client = MongoClient('localhost', 27017)
    db = client.online_shop
    prod_details = list(db.products.find({"_id": product_id}))
    return jsonify(prod_details)


@bp.route('/product/add', methods=('GET', 'POST'))
def prod_add():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        f = request.files

        if f['file'].filename == '':
            f = '/static/images/surface.jpg'
        else:
            f['file'].save(r'onlineshop/static/images/' + secure_filename(f['file'].filename))
            f = fr"/static/images/{secure_filename(f['file'].filename)}"

        prob_add = {"name_product": request.form.get('product_name'), "description": request.form.get('description'),
                    "category": request.form.get('category'),
                    "url_image": f}
        res = db.products.insert_one(prob_add)

        db = list(db.products.find({"_id": res.inserted_id}))
        for i in db:
            i["_id"] = str(i["_id"])

        return jsonify(db)


@bp.route('/product/edit', methods=('GET', 'POST'))
def prod_edit():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        f = request.files
        if f['file'].filename == '':
            url_image = list(db.products.find({"_id": ObjectId(request.form.get('_id'))}, {"url_image": 1, "_id": 0}))
        else:
            f['file'].save(r'onlineshop/static/images/' + secure_filename(f["file"].filename))
            url_image = [{"url_image": fr"/static/images/{secure_filename(f['file'].filename)}"}]

        db.products.update({"_id": ObjectId(request.form.get('_id'))}, {
            "$set": {"name_product": request.form.get('product_name'), "description": request.form.get('description'),
                     "category": request.form.get('category'),
                     "url_image": url_image[0]["url_image"]}})
        res = list(db.products.find({"_id": ObjectId(request.form.get('_id'))}))
        print(res)
        for i in res:
            i["_id"] = str(i["_id"])
        return jsonify(res)


@bp.route('/product/delete/<product_id>')
def prod_delete(product_id):
    client = MongoClient('localhost', 27017)
    db = client.online_shop
    db.products.delete_one({"_id": ObjectId(product_id)})
    res = list(db.products.find())
    for i in res:
        i["_id"] = str(i["_id"])
    return jsonify(res)


@bp.route('/product/upload', methods=('GET', 'POST'))
def upload_file_category():
    if request.method == 'POST' and request.files['file']:

        f = request.files['file']
        print(request.form)
        f.save(r'onlineshop/uploads/' + secure_filename(f.filename))
        try:
            category = check_validation_excel(fr'onlineshop/uploads/{secure_filename(f.filename)}')
            client = MongoClient('localhost', 27017)
            db = client.online_shop
            for i in category:
                res = db.products.insert_one(i)
                i['_id'] = str(res.inserted_id)
            return jsonify(category)
        except Exception as ex:
            return jsonify([{'error': f'{ex}'}])
        finally:
            os.remove(fr'onlineshop/uploads/{secure_filename(f.filename)}')
    return jsonify([{'error': "No File Added"}])


@bp.route('/product/json_category')
def prod_json_category():
    with io.open(r'onlineshop/category.json', encoding='utf-8') as f:
        f = load(f)
        return jsonify(f)


@bp.route('/inventory/list')
def inventory_list():
    client = MongoClient('localhost', 27017)
    db = client.online_shop
    inventory_list = list(db.inventory.find({}, {"name_inventory": 1}))
    for i in inventory_list:
        i["_id"] = str(i["_id"])
    return jsonify(inventory_list)


@bp.route('/inventory/delete/<inventory_id>')
def inventory_delete(inventory_id):
    client = MongoClient('localhost', 27017)
    db = client.online_shop
    db.inventory.delete_one({"_id": ObjectId(inventory_id)})
    res = list(db.inventory.find())
    for i in res:
        i["_id"] = str(i["_id"])
    return jsonify(res)


@bp.route('/inventory/add')
def inventory_add():
    client = MongoClient('localhost', 27017)
    db = client.online_shop
    count = db.inventory.count_documents({})
    new_inventory = {"name_inventory": f" انبار شماره {count + 1} ", "items": []}
    res = db.inventory.insert_one(new_inventory)
    new_inventory["_id"] = str(res.inserted_id)
    return jsonify([new_inventory])


@bp.route('/inventory/items', methods=('GET', 'POST'))
def inventory_items():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        data = request.form['_id']
        res = db.inventory.aggregate(
            [{"$unwind": {"path": "$items"}}, {"$match": {"_id": ObjectId(f"{data}")}},
             {"$project": {"items.name_product": 1, "_id": 0}}])
        return jsonify([{"name_product": i['items']['name_product']} for i in list(res)])


@bp.route('inventory/edit', methods=('GET', 'POST'))
def inventory_edit():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        if update(db, request):
            db.inventory.update({"_id": ObjectId(f'{request.form.get("_id")}'),
                                 "items": {"$elemMatch": {"name_product": f'{request.form.get("name_product")}'}}},
                                {"$set": {
                                    "items.$.quantity": int(request.form.get('quantity')),
                                    "items.$.price": int(request.form.get('price'))}})

            return jsonify([{"result": "update succsesfully"}, request.form])
        else:
            pass


@bp.route('inventory/add_prod', methods=('GET', 'POST'))
def inventory_add_prod():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        if update(db, request):
            db.inventory.update({"_id": ObjectId(f'{request.form.get("_id")}')},
                                {"$push": {
                                    "items": {
                                        "$each":
                                            [{"name_product": f'{request.form.get("name_product")}',
                                              "quantity": int(request.form.get('quantity')),
                                              "price": int(request.form.get('price')),
                                              "date_insert": datetime.now()}]}}})

        return jsonify([{"result": "update succsesfully"}, request.form])


@bp.route('inventory/edit_name_inventory', methods=('GET', 'POST'))
def inventory_edit_name_inventory():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        db.inventory.update({"_id": ObjectId(f'{request.form.get("_id")}')},
                            {"$set": {"name_inventory": request.form.get('name_inventory')}})

        return jsonify(
            [{"name_inventory": f"{request.form.get('name_inventory')}", "_id": f"{request.form.get('_id')}"}])


@bp.route('inventory/delete_prod', methods=('GET', 'POST'))
def inventory_delete_prod():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        db.inventory.update(
            {"_id": ObjectId(f'{request.form.get("_id")}')},
            {"$pull": {"items": {"name_product": request.form.get('name_product')}}}

        )

        return {'result': "delete action succsefully done ", "_id": request.form.get("_id")}


@bp.route('quantity/quantity_list')
def quantity_list():
    client = MongoClient('localhost', 27017)
    db = client.online_shop
    res = list(db.inventory.aggregate([{"$unwind": {"path": "$items"}}]))
    for i in res:
        i["_id"] = str(i["_id"])
        i["items"]["price"] = f'{i["items"]["price"]:,}'
    return jsonify(res)


@bp.route('quantity/quantity_add', methods=('GET', 'POST'))
def quantity_add():
    client = MongoClient('localhost', 27017)
    db = client.online_shop
    name_inventory = list(db.inventory.find({"_id": ObjectId(f'{request.form.get("_id")}')}, {"name_inventory": 1}))
    if update(db, request):
        res = {'result': False}
        db.inventory.update({"_id": ObjectId(f'{request.form.get("_id")}')},
                            {"$push": {
                                "items": {
                                    "$each":
                                        [{"name_product": f'{request.form.get("name_product")}',
                                          "quantity": int(request.form.get('quantity')),
                                          "price": int(request.form.get('price')),
                                          "date_insert": datetime.now()
                                          }]}}})
    else:
        res = {'result': True}

    for i in name_inventory:
        i["_id"] = str(i["_id"])

    res.update(name_inventory[0])
    res.update(request.form)
    return jsonify(res)


@bp.route('/order/list')
def order_list():
    client = MongoClient('localhost', 27017)
    db = client.online_shop

    res = list(db.basket.find())
    for i in res:
        i["_id"] = str(i["_id"])
        i["time_record"] = digits.en_to_fa(
            JalaliDate((JalaliDateTime.to_jalali(i["time_record"]))).strftime("%Y/%m/%d"))
        # i["total_costs"] = digits.en_to_fa(str(i["total_costs"]))
        i["total_costs"] = f"{i['total_costs']:,}"

    res = list(db.basket.find())
    for i in res:
        i["_id"] = str(i["_id"])
        i["time_record"] = digits.en_to_fa(
            JalaliDate((JalaliDateTime.to_jalali(i["time_record"]))).strftime("%Y/%m/%d"))
        # i["total_costs"] = digits.en_to_fa(str(i["total_costs"]))
        i["total_costs"] = f"{i['total_costs']:,}"

        i["time_give"] = digits.en_to_fa(
            JalaliDate((JalaliDateTime.to_jalali(i["time_give"]))).strftime("%Y/%m/%d"))
    return jsonify(list(res))


@bp.route('/basket/list')
def basket_list():
    client = MongoClient('localhost', 27017)
    db = client.online_shop

    x = session['basket']

    res = []
    for i in range(len(session.get('basket'))):
        y = list(db.inventory.aggregate(
            [{"$unwind": {"path": "$items"}}, {"$match": {"items.name_product": f"{x[i]['name_product']}"}},
             {"$sort": {"items.price": 1}},
             {"$limit": 1}]))
        res.extend(y)
        res[i].update({"quantity": f"{x[i]['quantity']}"})
    for i in res:
        i["_id"] = str(i["_id"])

    return jsonify(res)


@bp.route('/basket/delete', methods=('GET', 'POST'))
def basket_delete():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        res = db.inventory.update_many(
            {"_id": ObjectId(request.form.get('_id')), "items.name_product": request.form.get('name_product')},
            {"$inc": {"items.$.quantity": int(request.form.get('quantity'))}})

        session['basket'] = [i for i in session.get('basket') if i["name_product"] != f"{request.form['name_product']}"]
        print(session.get('basket'))
        return request.form


@bp.route('/basket/record_form', methods=('GET', 'POST'))
def record_form():
    if request.method == "POST":
        client = MongoClient('localhost', 27017)
        db = client.online_shop
        items = session['basket']

        basket = {
            "customer_first_name": request.form['name'],
            "customer_last_name": request.form['family_name'],
            "time_give": JalaliDateTime.fromisoformat(request.form['date-give']).to_gregorian(),
            "time_record": datetime.now(),
            "phone": int(request.form['phone']),
            "total_costs": int(request.form['totalCoast']),
            "address": request.form['address'],
            "items": []
        }
        products = {i: request.form[i].split(',') for i in request.form if i.startswith('product')}
        basket["items"] = [
            dict({"name_product": products[i][0], "quantity": int(products[i][1]), "price": int(products[i][2]),
                  "name_inventory": products[i][3]})
            for i in products]
        db.basket.insert_one(basket)
        session['basket'] = []
        return request.form


@bp.route('/add_to_basket', methods=('GET', 'POST'))
def add_to_basket():
    basket = session.get('basket')
    if request.method == "POST":
        add = False
        if basket is None:
            # basket = []
            session['basket'] = []
        for product in session['basket']:
            if product["name_product"] == request.form['name_product']:
                product["quantity"] = int(product["quantity"]) + int(request.form['quantity'])
                add = True
                break
        if add is False:
            session.get('basket').append(
                {"name_product": request.form["name_product"], "quantity": request.form["quantity"]})

        client = MongoClient('localhost', 27017)
        db = client.online_shop
        update_reverse(db, request)
        session['bug'] = None
        return {"result": "کالای مورد نظر به سبد خرید شما اضافه شد"}
